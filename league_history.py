# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import openpyxl
import pandas as pd
import numpy as np
import os, sys
import warnings
warnings.filterwarnings("ignore")
if 'input' in os.listdir(os.getcwd()):
    INPUT_FILE_PATH = os.getcwd()
elif 'input' in os.listdir(os.path.join(os.getcwd(),'..','..')):
    INPUT_FILE_PATH = os.path.join(os.getcwd(),'..','..')
else:
    INPUT_FILE_PATH = os.getcwd()

def get_draft_slots(data,position,player):
    return data.loc[(data['position']==position) & (data['player'].str.lower().str.contains(player.lower()))].sort_values(by='Year',ascending = False)

def load_performance_data():
    stats = pd.read_excel(os.path.join(INPUT_FILE_PATH,'input','league_performance.xlsx'))
    stats['POINTS OVER MEAN'] = stats['PF/G']-stats.groupby('YEAR')['PF/G'].transform('mean')
    stats['MONIKOR'] = stats['YEAR'].astype(str)+stats['PLAYER']
    return stats

def get_performance_stats(data, subset,stats):
    monikors = subset['MONIKOR'].unique()
    percentage_teams = len(monikors)/len(data['MONIKOR'].unique())
    stats_subset = stats.loc[stats['MONIKOR'].isin(monikors)]
    stats_agg = stats_subset.agg({'POINTS OVER MEAN':'mean','W':'sum','L':'sum'})
    stats_agg.loc['W/L %'] = stats_agg.loc['W']/(stats_agg.loc['W'] + stats_agg.loc['L'])
    stats_agg.loc['% of TEAMS'] = percentage_teams
    return stats_agg

def get_all_historical_pick_probabilities(data, buffer = 2, weight_length = 3 ):
    data_use = data.copy()
    n = len(data_use['Year'].unique())
    weights = [np.exp(-k/weight_length) for k in np.arange(n)]
    weights = [weight/sum(weights) for weight in weights]
    weights_df = pd.DataFrame(weights,index = sorted(list(data_use['Year'].unique()),reverse = True)).reset_index().rename(columns = {0:'Weights','index':'Year'})
    data_use = pd.merge(data_use, weights_df, on = ['Year'],how = 'left')
    final_perc = pd.DataFrame(columns = ['QB','RB','WR','TE'], index = np.arange(1,12*17+1))
    for pick in np.arange(1,12*17+1):
        slice = data_use.loc[(data_use['pick_type']=='PICK') & (data_use['pick'] <= pick+buffer) & (data_use['pick'] >= pick - buffer)]
        slice_agg1 = slice.groupby(by = ['Year','position']).agg({'rank':'count'}).reset_index()
        slice_agg2 = slice.groupby(by = ['Year']).agg({'rank':'count','Weights':'mean'}).reset_index()
        slice_agg2['wgt'] = slice_agg2['Weights']/slice_agg2['rank']
        slice_agg = pd.merge(slice_agg1,slice_agg2[['Year','wgt']],on = 'Year',how = 'left')
        slice_agg['perc'] = slice_agg['rank']*slice_agg['wgt']
        answer = slice_agg[['position','perc']].groupby(by = 'position').sum()
        answer = answer.reindex(index = ['QB','RB','WR','TE']).fillna(0)
        final_perc.loc[pick] = answer['perc']
    return final_perc

def get_personal_pick_probability_by_round(data,user,round,weight_length = 3):
    data_use = data.copy()
    n = len(data_use['Year'].unique())
    weights = [np.exp(-k/weight_length) for k in np.arange(n)]
    weights = [weight/sum(weights) for weight in weights]
    weights_df = pd.DataFrame(weights,index = sorted(list(data_use['Year'].unique()),reverse = True)).reset_index().rename(columns = {0:'Weights','index':'Year'})
    data_use = pd.merge(data_use, weights_df, on = ['Year'],how = 'left')
    slice = data_use.loc[(data_use['pick_type']=='PICK') & (data_use['round'] == round) & (data_use['Player'] == user)]
    slice_agg1 = slice.groupby(by=['Year', 'position']).agg({'rank': 'count'}).reset_index()
    slice_agg2 = slice.groupby(by=['Year']).agg({'rank': 'count', 'Weights': 'mean'}).reset_index()
    slice_agg2['wgt'] = slice_agg2['Weights'] / slice_agg2['rank']
    slice_agg2['wgt']/=slice_agg2['wgt'].sum()
    slice_agg = pd.merge(slice_agg1, slice_agg2[['Year', 'wgt']], on='Year', how='left')
    slice_agg['perc'] = slice_agg['rank'] * slice_agg['wgt']
    answer = slice_agg[['position', 'perc']].groupby(by='position').sum()
    answer = answer.reindex(index=['QB', 'RB', 'WR', 'TE']).fillna(0)
    return answer

def parse_historic_data():
    draft_excel_file = "BKFFL Draft History.xlsx"
    wb = openpyxl.load_workbook(os.path.join(INPUT_FILE_PATH,'input',draft_excel_file))
    tabs_dict = {
        '2022': {
            'row_offset':0, 'col_offset':0, 'num_rounds': 17, 'num_teams': 12
        },
        '2021': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2020': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2019': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2018': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2017': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2016': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 14
        },
        '2015': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 14
        }
    }
    draft_dict = {}
    for year, tab_dict in tabs_dict.items():
        tab_name = year + ' Draft'
        num_teams = tab_dict['num_teams']
        num_rounds = tab_dict['num_rounds']
        ws = wb[tab_name]
        cell_color_dict = {}
        for col in range(2,2+num_teams+tab_dict['col_offset']):
            cell = ws.cell(1+tab_dict['row_offset'],col)
            cell_text = cell.value
            cell_color = cell.fill.start_color.rgb
            cell_color_dict[cell_color] = {'player':cell_text, 'team':[]}
        reject_list = []
        for row in range(2+tab_dict['row_offset'], num_rounds+2+tab_dict['row_offset']):
            for col in range(2+tab_dict['col_offset'], num_teams+2+tab_dict['col_offset']):
                round = row-1-tab_dict['row_offset']
                pick_in_round = col - 1 -tab_dict['col_offset']
                if round%2==0:
                    pick_in_round = num_teams+1-pick_in_round
                pick = (round-1)*num_teams+pick_in_round
                cell = ws.cell(row,col)
                cell_color = cell.fill.start_color.rgb
                cell_text = cell.value
                pick_type = 'PICK'
                if cell.font.bold:
                    pick_type = 'KEEP'
                position = cell_text.split(' ',1)[0]
                player = cell_text.split(' ',1)[1]

                if cell_color in cell_color_dict.keys():
                    cell_color_dict[cell_color]['team'] += [(position,player,pick_type,round,pick_in_round, pick)]
                else:
                    reject_list += [cell_text]
                    print("{},{},{}".format(year,row,col))
        # Fix keys
        data_dict = {}
        def custom_sort_key(item):
            return item[5]
        for key, subdict in cell_color_dict.items():
            player_name = subdict['player']
            subdict['color'] = key
            subdict['team'] = sorted(subdict['team'],key = custom_sort_key)
            subdict['num_players'] = len(subdict['team'])
            data_dict[player_name] = subdict
        draft_dict[year] = data_dict
    year = '2019'
    draft_dict[year]['Paskal']
    # for year, subdict in draft_dict.items():
    #     for player, subsubdict in subdict.items():
    #         print("{}\t{}\t{}".format(year,player,subsubdict['num_players']))

    # Turn into dict for each person and get all players
    person_dict = {}
    data = pd.DataFrame()
    columns = ['position','player','pick_type','round','pick_in_round','pick']
    for year, subdict in draft_dict.items():
        for player, subsubdict in subdict.items():
            if player not in person_dict:
                person_dict[player] = {}
            person_dict[player][year] = subsubdict.copy()
            team_info = pd.DataFrame(subsubdict['team'],columns = columns)
            team_info.insert(0,'Year',int(year))
            team_info.insert(1,'Player',player)
            data = pd.concat([data,team_info])
    data = data.sort_values(by = ['Year','pick'],ascending = [False,True]).reset_index(drop = True)
    data.loc[data['player'] == 'D/ST','player'] = data.loc[data['player'] == 'D/ST','position']
    data.loc[~data['position'].isin(['QB','RB','WR','TE','D/ST']),'position']='D/ST'
    data.loc[data['pick_type']=='PICK','rank'] = data.loc[data['pick_type']=='PICK'].groupby(['Year','position'])['pick'].rank()
    data['rank'] = data['rank'].fillna(1000).astype(int)
    data.loc[data['pick_type'] == 'PICK', 'players_pick'] = data.loc[data['pick_type'] == 'PICK'].groupby(['Year', 'Player'])['pick'].rank()
    data['players_pick'] = data['players_pick'].fillna(1000).astype(int)
    data['MONIKOR']=data['Year'].astype(str)+data['Player']
    return data


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    draft_excel_file = "BKFFL Draft History.xlsx"
    wb = openpyxl.load_workbook(os.path.join(INPUT_FILE_PATH,'input',draft_excel_file))
    tabs_dict = {
        '2022': {
            'row_offset':0, 'col_offset':0, 'num_rounds': 17, 'num_teams': 12
        },
        '2021': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2020': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2019': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2018': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2017': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 12
        },
        '2016': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 14
        },
        '2015': {
            'row_offset': 0, 'col_offset': 0, 'num_rounds': 17, 'num_teams': 14
        }
    }
    draft_dict = {}
    for year, tab_dict in tabs_dict.items():
        tab_name = year + ' Draft'
        num_teams = tab_dict['num_teams']
        num_rounds = tab_dict['num_rounds']
        ws = wb[tab_name]
        cell_color_dict = {}
        for col in range(2,2+num_teams+tab_dict['col_offset']):
            cell = ws.cell(1+tab_dict['row_offset'],col)
            cell_text = cell.value
            cell_color = cell.fill.start_color.rgb
            cell_color_dict[cell_color] = {'player':cell_text, 'team':[]}
        reject_list = []
        for row in range(2+tab_dict['row_offset'], num_rounds+2+tab_dict['row_offset']):
            for col in range(2+tab_dict['col_offset'], num_teams+2+tab_dict['col_offset']):
                round = row-1-tab_dict['row_offset']
                pick_in_round = col - 1 -tab_dict['col_offset']
                if round%2==0:
                    pick_in_round = num_teams+1-pick_in_round
                pick = (round-1)*num_teams+pick_in_round
                cell = ws.cell(row,col)
                cell_color = cell.fill.start_color.rgb
                cell_text = cell.value
                pick_type = 'PICK'
                if cell.font.bold:
                    pick_type = 'KEEP'
                position = cell_text.split(' ',1)[0]
                player = cell_text.split(' ',1)[1]

                if cell_color in cell_color_dict.keys():
                    cell_color_dict[cell_color]['team'] += [(position,player,pick_type,round,pick_in_round, pick)]
                else:
                    reject_list += [cell_text]
                    print("{},{},{}".format(year,row,col))
        # Fix keys
        data_dict = {}
        def custom_sort_key(item):
            return item[5]
        for key, subdict in cell_color_dict.items():
            player_name = subdict['player']
            subdict['color'] = key
            subdict['team'] = sorted(subdict['team'],key = custom_sort_key)
            subdict['num_players'] = len(subdict['team'])
            data_dict[player_name] = subdict
        draft_dict[year] = data_dict
    year = '2019'
    draft_dict[year]['Paskal']
    # for year, subdict in draft_dict.items():
    #     for player, subsubdict in subdict.items():
    #         print("{}\t{}\t{}".format(year,player,subsubdict['num_players']))

    # Turn into dict for each person and get all players
    person_dict = {}
    data = pd.DataFrame()
    columns = ['position','player','pick_type','round','pick_in_round','pick']
    for year, subdict in draft_dict.items():
        for player, subsubdict in subdict.items():
            if player not in person_dict:
                person_dict[player] = {}
            person_dict[player][year] = subsubdict.copy()
            team_info = pd.DataFrame(subsubdict['team'],columns = columns)
            team_info.insert(0,'Year',int(year))
            team_info.insert(1,'Player',player)
            data = pd.concat([data,team_info])
    data = data.sort_values(by = ['Year','pick'],ascending = [False,True]).reset_index(drop = True)
    data.loc[data['player'] == 'D/ST','player'] = data.loc[data['player'] == 'D/ST','position']
    data.loc[~data['position'].isin(['QB','RB','WR','TE','D/ST']),'position']='D/ST'
    data.loc[data['pick_type']=='PICK','rank'] = data.loc[data['pick_type']=='PICK'].groupby(['Year','position'])['pick'].rank()
    data['rank'] = data['rank'].fillna(1000).astype(int)
    data.loc[data['pick_type'] == 'PICK', 'players_pick'] = data.loc[data['pick_type'] == 'PICK'].groupby(['Year', 'Player'])['pick'].rank()
    data['players_pick'] = data['players_pick'].fillna(1000).astype(int)
    data['MONIKOR']=data['Year'].astype(str)+data['Player']


    stats = load_performance_data()

    subset = get_draft_slots(data,'TE','Kelce')
    get_performance_stats(data,subset,stats)

    subset = data.loc[(data['pick_type'] == 'PICK') & (data['round'] == 1) & (data['position']=='WR')]
    get_performance_stats(data,subset,stats)

    subset = data.loc[(data['pick_type'] == 'PICK') & (data['round'].isin([1,2,3])) & (data['position']=='TE')]
    get_performance_stats(data, subset, stats)

    subset = data.loc[(data['pick_type'] == 'PICK') & (data['rank'].isin([1,2,3])) & (data['position']=='QB')]
    get_performance_stats(data, subset, stats)

    subset = data.loc[(data['player'].str.contains('A. Rodgers')) &(data['position']=='QB')]
    get_performance_stats(data, subset, stats)

    # Get player stats
    player_stats = stats[['PLAYER','W','L','POINTS OVER MEAN','CHAMP POINTS']].groupby(by = 'PLAYER')\
        .agg({'W':'sum','L':'sum','POINTS OVER MEAN':'mean','CHAMP POINTS':'sum'})
    player_stats['WIN %'] = player_stats['W']/(player_stats['W'] + player_stats['L'])
    player_stats = player_stats.sort_values(by = 'WIN %',ascending = False)
    players_now = stats.loc[stats['YEAR'] == 2022,'PLAYER'].unique()
    player_stats = player_stats.loc[player_stats.index.isin(players_now)]

    data = pd.merge(data,stats[['MONIKOR','W','L']],on='MONIKOR',how = 'left')


    def concat_players(x, pick_values):
        positions =  x.loc[x['players_pick'].isin(sorted(pick_values))].sort_values('players_pick')['position']
        return ','.join(positions)
    first_2_picks = [1,2]
    first_3_picks = [1,2,3]
    first_5_picks = [1,2,3,4,5]
    first_2 = pd.DataFrame(data.loc[data['players_pick'].isin(first_2_picks)].groupby(['MONIKOR']).apply(concat_players,pick_values=first_2_picks).reset_index()).rename(
        columns={0: 'FIRST_2'})
    first_3 = pd.DataFrame(data.loc[data['players_pick'].isin(first_3_picks)].groupby(['MONIKOR']).apply(concat_players, pick_values = first_3_picks).reset_index()).rename(columns= {0:'FIRST_3'})
    first_5 = pd.DataFrame(data.loc[data['players_pick'].isin(first_5_picks)].groupby(['MONIKOR']).apply(concat_players, pick_values = first_5_picks).reset_index()).rename(columns= {0:'FIRST_5'})
    data = pd.merge(data, first_2, on='MONIKOR', how='left')
    data = pd.merge(data,first_3,on = 'MONIKOR', how = 'left')
    data = pd.merge(data,first_5,on = 'MONIKOR',how = 'left')

    draft_stats = data[['Year','Player','W','L','FIRST_2','FIRST_3','FIRST_5']].drop_duplicates()
    draft_stats_2_agg = draft_stats.groupby('FIRST_2').agg({'Year':'mean','W':'sum','L':'sum'})
    draft_stats_2_agg['W/L %'] = draft_stats_2_agg['W']/(draft_stats_2_agg['W']+draft_stats_2_agg['L'])
    draft_stats_2_agg = draft_stats_2_agg.sort_values(by = 'W/L %',ascending = False)
    draft_stats_3_agg = draft_stats.groupby('FIRST_3').agg({'Year':'mean','W':'sum','L':'sum'})
    draft_stats_3_agg['W/L %'] = draft_stats_3_agg['W']/(draft_stats_3_agg['W']+draft_stats_3_agg['L'])
    draft_stats_3_agg = draft_stats_3_agg.sort_values(by = 'W/L %',ascending = False)
    draft_stats_5_agg = draft_stats.groupby('FIRST_5').agg({'Year':'mean','W':'sum','L':'sum'})
    draft_stats_5_agg['W/L %'] = draft_stats_5_agg['W']/(draft_stats_5_agg['W']+draft_stats_5_agg['L'])
    draft_stats_5_agg = draft_stats_5_agg.sort_values(by = 'W/L %',ascending = False)
    # data['first_3'] = data.groupby(['MONIKOR'])['position'].agg('#'.join)

