import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import league_history, download_data
import time
import os, sys
import ast
import datetime
import shutil

if 'input' in os.listdir(os.getcwd()):
    INPUT_FILE_PATH = os.getcwd()
elif 'input' in os.listdir(os.path.join(os.getcwd(),'..','..')):
    INPUT_FILE_PATH = os.path.join(os.getcwd(),'..','..')
else:
    INPUT_FILE_PATH = os.getcwd()

rookie_list = {
    'Anthony Richardson', 'Bryce Young','C.J. Stroud','Will Levis','Clayton Tune','Bijan Robinson','Jahmyr Gibbs',
    'Zach Charbonnet','De\'Von Achane','Kendre Miller','Tank Bigsby','Roschon Johnson','Tyjae Spears','Jordan Addison','Jaxon Smith-Njigba',
    'Quentin Johnston','Zay Flowers','Dalton Kincaid','Michael Mayer','Luke Musgrave','Luke Schoonmaker','Darnell Washington',
    'Israel Abanikanda','DeWayne McBride','Sean Tucker','Jonathan Mingo','Rashee Rice','Jalin Hyatt','Jayden Reed','Josh Downs','Marvin Mims Jr.','Tank Dell'

}


def construct_fp_rankings():
    generated_player_ids = []
    def generate_player_id(row):
        position = row['Position']

        player_count = position_sequence[position]
        while True:
            # Generate a random number between 1000 and 9999
            random_number = random.randint(100, 999)

            player_id = f"{player_count}{random_number:03d}"

            # Check if the generated ID already exists in the list of generated IDs
            if player_id not in generated_player_ids:
                generated_player_ids.append(player_id)
                return player_id
        return player_id

    random.seed(42)
    try:
        fp_rankings = download_data.get_fantasy_pros_rankings_data()
        print("Download successful.")
    except:
        fp_rankings = pd.read_csv(os.path.join(INPUT_FILE_PATH,'input','FantasyPros_2023_Draft_OP_Rankings.csv'))
        print("Rankings download failed. Loading default (hard-saved) rankings.")
    position_sequence = {'QB': 1, 'RB': 2, 'WR': 3, 'TE': 4}
    fp_rankings = fp_rankings.loc[fp_rankings['POS'].str[:2].isin(position_sequence.keys())]
    fp_rankings = fp_rankings.drop(columns=['TIERS', 'BEST', 'WORST']).rename(
        columns={'RK': 'Rank', 'PLAYER NAME': 'Player', 'TEAM': 'Team',
                 'POS': 'Position', 'AVG.': 'Avg', 'STD.DEV': 'Sigma', 'ECR VS. ADP': 'ADP'})
    fp_rankings.loc[fp_rankings['ADP'] == '-', 'ADP'] = 0
    fp_rankings['ADP'] = fp_rankings['ADP'].astype(int)
    fp_rankings['ADP'] += fp_rankings['Rank']
    fp_rankings = fp_rankings.reset_index(drop=True)
    fp_rankings['Rank'] = fp_rankings.index + 1
    fp_rankings = fp_rankings.set_index('Rank')
    fp_rankings['Position'] = fp_rankings['Position'].str[:2]
    fp_rankings['ID'] = fp_rankings.apply(generate_player_id, axis=1)
    return fp_rankings

def shuffle_players_locking_qbs(rankings_old, rankings_new):
    rankings = rankings_old.reset_index(drop = True)
    rankings.index +=1
    rankings_new_qb = rankings_new.loc[rankings_new['Position']=='QB']
    rankings.loc[rankings['Position']=='QB','ID'].iloc[:len(rankings_new_qb)]= rankings_new_qb['ID']
    rankings = pd.merge(rankings_new, rankings.drop(columns = ['Player','Team']),on = 'ID',how = 'left')


def convert_to_position(rankings_all):
    position_list = ['QB','RB','WR','TE']
    position_rank = {}
    for position in position_list:
        df = rankings_all.loc[rankings_all['Position'] == position].reset_index()
        df.index = df.index + 1
        df.index.name = 'Positional Rank'
        position_rank[position] = df
    return position_rank

def remove_keepers(fp_rankings, keepers):
    keepers_use = keepers.copy()
    if 'ID' in keepers_use.columns:
        keepers_use = keepers_use.drop(columns = ['ID'])
    rankings_use = fp_rankings.copy()
    rankings_use['Player_full'] = rankings_use['Player']
    rankings_use['Player'] = rankings_use['Player'].str.split(n=2).str[:2].str.join(' ')
    rankings_use['Player'] = rankings_use['Player'].str.split(' ',1).str[0].str[0] + '. ' + rankings_use['Player'].str.split(' ',1).str[1]
    keepers_use = pd.merge(keepers_use.reset_index(), rankings_use[['Position','Player','Player_full','ID']],
                           on = ['Position','Player'],how = 'left').set_index('Pick')

    # OVERRIDE - confusion Bijan vs Brian Robinson
    players_to_drop = [        'Bijan Robinson'  ]
    keepers_use = keepers_use.loc[~keepers_use['Player_full'].isin(players_to_drop)].drop(columns= 'Player_full')

    new_rankings = fp_rankings.loc[~fp_rankings['ID'].isin(keepers_use['ID'].unique())]
    return new_rankings, keepers_use


def create_random_rankings(fp_rankings, bump_rookies = True):
    new_rank = fp_rankings.copy()
    new_rank['rank_copy'] = new_rank.index
    new_rank['gauss']=np.random.randn(len(new_rank))
    new_rank['gauss'] = new_rank['gauss'].apply(lambda x: max(-1,x))
    new_rank['up_fac'] = new_rank['rank_copy'].apply(lambda x: min(1,.8/50.*x+.2))
    new_rank['down_fac'] = new_rank['rank_copy'].apply(lambda x: min(1,.8/50.*x+.1))
    if bump_rookies:
        new_rank.loc[(new_rank['Player'].isin(rookie_list)) & (new_rank.index >= 30),'down_fac']*=1.5
        new_rank.loc[new_rank['Player'].isin(rookie_list), 'up_fac']*=0
    # Tweek TEs
    new_rank.loc[(new_rank['Position']=='TE') & (new_rank['Player']!= 'Travis Kelce'),'down_fac']*=1
    new_rank.loc[(new_rank['Position']=='TE') & (new_rank['Player']!= 'Travis Kelce'),'up_fac']*=2

    new_rank.loc[new_rank['gauss'] < 0,'gauss']*= new_rank.loc[new_rank['gauss'] < 0,'down_fac']
    new_rank.loc[new_rank['gauss'] >= 0,'gauss']*= new_rank.loc[new_rank['gauss'] >= 0,'up_fac']
    new_rank['rank_new'] = new_rank['Avg'] + new_rank['Sigma']*new_rank['gauss']
    new_rank = new_rank.sort_values(by = 'rank_new', ascending = True)

    new_rank.index = new_rank.reset_index().index + 1
    new_rank = new_rank.drop(columns = ['gauss','rank_new','up_fac','down_fac','rank_copy'])
    new_rank.index.name = 'Rank'
    new_position = convert_to_position(new_rank)
    return new_rank, new_position

def load_current_setup(file_name, rankings):

    num_rounds = 17
    num_teams = 12
    wb = openpyxl.load_workbook(file_name)
    year = '2023'

    tab_name = year + ' Draft'
    ws = wb[tab_name]
    cell_color_dict = {}
    pick_dict = {}
    for col in range(2,2+num_teams):
        cell = ws.cell(1,col)
        cell_text = cell.value
        cell_color = cell.fill.start_color.rgb
        # print("{}\t{}".format(cell_text,cell_color))
        cell_color_dict[cell_color] = cell_text
        pick_dict[cell_text] = col - 1
    reject_list = []
    keepers = []

    draft_record = []
    for row in range(2, num_rounds+2):
        for col in range(2, num_teams+2):
            round = row-1
            slot = col - 1
            if round%2==0:
                slot = num_teams+1-slot
            pick = (round-1)*num_teams+slot
            cell = ws.cell(row,col)
            cell_color = cell.fill.start_color.rgb
            cell_text = str(cell.value)
            if cell_color in cell_color_dict.keys():
                owner = cell_color_dict[cell_color]
                if cell.font.bold:
                    position = cell_text.split(' ', 1)[0]
                    player = cell_text.split(' ', 1)[1]
                    pick_type = 'KEEP'
                    draft_record += [(owner, round, slot, pick, pick_type, position, player)]
                elif str(cell_text).isnumeric():
                    pick_type = 'PICK'
                    draft_record += [(owner, round, slot, pick, pick_type, '', '')]
                else:
                    print("Parse Error\t{},{},{}".format(year, row, col))
            else:
                reject_list += [cell_text]
                print("{},{},{}".format(year, row, col))
    draft_record = pd.DataFrame(draft_record, columns = ['Owner','Round','Slot','Pick','Pick Type', 'Position','Player']).set_index('Pick')
    keepers = draft_record.loc[draft_record['Pick Type'] == 'KEEP']
    _, keepers = remove_keepers(rankings, keepers)
    draft_record = pd.merge(draft_record.reset_index(), keepers[['Position','Player','ID']], on = ['Position','Player'],how = 'left').fillna('').set_index('Pick').sort_index()
    return draft_record, keepers, pick_dict


class MockDraft:
    def __init__(self):
        self.historic_data = league_history.parse_historic_data()
        self.historic_pick_probs = league_history.get_all_historical_pick_probabilities(self.historic_data,buffer = 2, weight_length = 3)
        self.fp_rankings = construct_fp_rankings()
        self.projections = self.load_projections()
        self.initial_draft, keepers, pick_dict = load_current_setup(    file_name = os.path.join(INPUT_FILE_PATH,'input',"BKFFL Draft History.xlsx"), rankings = self.fp_rankings)
        self.pick_order = pick_dict.keys()
        _, self.keepers = remove_keepers(self.fp_rankings,keepers)
        random.seed()
        self.fp_position_ranks = convert_to_position(self.fp_rankings)
        self.player_dict = {}
        for owner in self.initial_draft['Owner'].unique():
            self.player_dict[owner] = {}
            self.player_dict[owner]['ranks'], self.player_dict[owner]['pos_ranks'] = create_random_rankings(self.fp_rankings)
        self.draft_board = self.initial_draft.copy()
        self.user_name_list = sorted(self.player_dict.keys())

        short_rankings = self.fp_rankings.copy()
        short_rankings['Player (short)'] = short_rankings['Player'].str.split(' ',1).str[0].str[0] + '. ' \
                                           + short_rankings['Player'].str.split(' ',1).str[1]
        self.id_mapping = short_rankings[['ID','Player (short)']].set_index('ID')['Player (short)'].to_dict()
        self.id_pos_mapping = short_rankings[['ID','Position']].set_index('ID')['Position'].to_dict()
        self.replacement_ids = self.load_replacement_ids()

        adp_half = download_data.get_fantasy_pros_adp_data(use_ppr = False)
        adp_full = download_data.get_fantasy_pros_adp_data(use_ppr = True)
        adp = pd.merge(adp_half[['PLAYER','POS','YAHOO','SLEEPER','RTSPORTS']],adp_full[['PLAYER','POS','ESPN','NFL']], on = ['PLAYER','POS'])
        adp['PLAYER'] = adp['PLAYER'].replace({'Dalvin Cook':'Dalvin Cook d ()','Kareem Hunt':'Kareem Hunt d ()','Leonard Fournette':'Leonard Fournette d ()',
                                           'Ezekiel Elliott':'Ezekiel Elliott d ()','Jarvis Landry':'Jarvis Landry d ()'})

        def average_non_empty(row):
            non_empty_values = pd.to_numeric(row, errors='coerce').dropna()
            if not non_empty_values.empty:
                return non_empty_values.mean()
            else:
                return None
        adp['AVG'] = adp[['YAHOO','SLEEPER','RTSPORTS','ESPN','NFL']].apply(average_non_empty, axis=1)
        adp = adp.sort_values(by = 'AVG').drop(columns = 'AVG').reset_index(drop = True)
        adp.index +=1
        adp['PLAYER'] = adp['PLAYER'].str.split(' ', 5).str[:-2].str.join(' ')
        rankings = self.fp_rankings[['Player','Position','ID']]
        adp = pd.merge(rankings, adp.rename(columns = {'PLAYER':'Player','POS':'Position'}), on=['Player','Position'], how='right')




    def get_id(self,player_name):
        try:
            return self.fp_rankings.set_index('Player').loc[player_name,'ID']
        except:
            print("No player ID found. Returning null.")
            return ''

    def reset_player_draft_rankings(self):
        player_dict = {}
        for owner in self.initial_draft['Owner'].unique():
            player_dict[owner] = {}
            player_dict[owner]['ranks'], self.player_dict[owner]['pos_ranks'] = create_random_rankings(self.fp_rankings)
        return player_dict

    def load_replacement_ids(self):
        replacements = {'QB1':'Derek Carr','OP':'Ryan Tannehill','RB1':'Cam Akers','RB2':'Samaje Perine','WR1':'Christian Kirk','WR2':'Jahan Dotson','Flex1':'Rashod Bateman',
                        'Flex2':'Courtland Sutton','TE1':'Dawson Knox'}
        replacement_id = {key:self.get_id(val) for key, val in replacements.items()}
        return replacement_id

    def load_projections(self):
        def compute_100_yd_games(yards):
            a = 11/3840000
            b = 1 - 11*160/3840
            return max(0,a*yards**2 + b)
        def compute_300_yd_pass_games(yards):
            a = 11 / 3840000
            b = 1 - 11 * 160 / 3840
            return max(0, a * (yards/3) ** 2 + b)

        # Parse Mike Clay projections
        file_name_espn = 'ESPN_projections_20230806.csv'
        df = pd.read_csv(os.path.join(INPUT_FILE_PATH,'input',file_name_espn))
        col_rename =  {'Pass Yds': ('Pass', 'YDS'), 'Pass TD': ('Pass', 'TDS'), 'Pass INT': ('Pass', 'INTS'),
             'Rush YDS': ('Rush', 'YDS'), 'Rush TD': ('Rush', 'TDS'), 'REC': ('Rec', 'REC'), 'REC YDS': ('Rec', 'YDS'),
             'REC TDS': ('Rec', 'TDS'),'Rank':('','Rank'),'Player':('','Player'),'GAMES':('','GAMES')}
        df = df.rename(columns = col_rename)
        df.columns = pd.MultiIndex.from_tuples(df.columns)
        for col in df.columns:
            try:
                df[col] = df[col].str.replace('-','0').fillna(0).astype(float)
            except:
                pass
        df.insert(5, ('Pass', '300 GMS'), df[('Pass', 'YDS')].apply(compute_300_yd_pass_games).round(1))
        df.insert(8, ('Rush', '100 GMS'), df[('Rush', 'YDS')].apply(compute_100_yd_games).round(1))
        df.insert(12, ('Rec', '100 GMS'), df[('Rec', 'YDS')].apply(compute_100_yd_games).round(1))
        ffp = df[('Pass', 'YDS')] / 25 + df[('Pass', 'TDS')] * 4 + df[('Pass', 'INTS')] * (-2) + df[
            ('Pass', '300 GMS')] * 2 \
              + df[('Rush', 'YDS')] / 10 + df[('Rush', 'TDS')] * 6 + df[
                  ('Rush', '100 GMS')] * 2 \
              + df[('Rec', 'REC')] * 0.5 + df[('Rec', 'YDS')] / 10 + df[('Rec', 'TDS')] * 6 + df[
                  ('Rec', '100 GMS')] * 2
        df.insert(2, ('', 'FFP'), ffp.round(1))
        df.insert(3, ('', 'FPPG'), (ffp / df['']['GAMES'].apply(lambda x:max(x,1))).round(1))
        df_clay = df.copy()
        rankings = self.fp_rankings[['Player','Position','ID']].rename(columns = {'Player':('','Player'),'Position':('','POS'),'ID':('','ID')})
        df_clay = pd.merge(rankings, df_clay, on=[('', 'Player')], how='right')
        df_clay.columns = pd.MultiIndex.from_tuples(df_clay.columns)


        # file_name_qb = 'FantasyPros_Fantasy_Football_Projections_QB.csv'
        # df_qb = pd.read_csv(os.path.join(INPUT_FILE_PATH,'input',file_name_qb)).dropna(subset = ['Team']).drop(columns = ['ATT','CMP','FPTS', 'ATT.1'])
        # df_qb['Player'] = df_qb['Player'].fillna(method = 'ffill')
        # df_qb.loc[~df_qb['Team'].isin(['high','low']),'Team'] = 'median'
        # df_qb = df_qb.rename(columns = {'Player':('','Player'),'Team':('','Team'),'YDS':('Pass','YDS'),'TDS':('Pass','TDS'),
        #                                 'INTS':('Pass','INTS'),'YDS.1':('Rush','YDS'),'TDS.1':('Rush','TDS'),'FL':('Rush','FL')})
        # df_qb[('Pass','YDS')] = df_qb[('Pass','YDS')].str.replace(',','').astype(float)
        # df_qb[('Rush','YDS')] = df_qb[('Rush','YDS')].str.replace(',','').astype(float)
        # df_qb.columns = pd.MultiIndex.from_tuples(df_qb.columns)
        # df_qb.insert(5,('Pass','300 GMS'),df_qb[('Pass','YDS')].apply(compute_300_yd_pass_games).round(1))
        # df_qb.insert(9,('Rush','100 GMS'),df_qb[('Rush','YDS')].apply(compute_100_yd_games).round(1))
        # df_qb.insert(2,('','POS'),'QB')


        # # Parse FantasyPros projections.
        # file_name_flex = 'FantasyPros_Fantasy_Football_Projections_FLX.csv'
        # df_flex = pd.read_csv(os.path.join(INPUT_FILE_PATH,'input',file_name_flex)).dropna(subset = ['Team'])
        # # realign the columns
        # df_flex_mis = df_flex.loc[df_flex['Player'].isna()]
        # df_flex_mis.iloc[:,3:]=df_flex_mis.iloc[:,2:-1]
        # df_flex_mis['POS'] = np.nan
        # df_flex.loc[df_flex['Player'].isna()] = df_flex_mis
        # # end realign
        # df_flex['Player'] = df_flex['Player'].fillna(method = 'ffill')
        # df_flex['POS'] = df_flex['POS'].fillna(method = 'ffill')
        # df_flex['POS'] = df_flex['POS'].str[:2]
        # df_flex = df_flex.loc[df_flex['POS'].isin(['QB','RB','WR','TE'])].reset_index(drop = True)
        # df_flex.loc[~df_flex['Team'].isin(['high','low']),'Team'] = 'median'
        # df_flex = df_flex.drop(columns = ['ATT','FPTS']).rename(columns = {'Player':('','Player'),'Team':('','Team'),'POS':('','POS'),
        #                                                                    'YDS':('Rush','YDS'),'TDS':('Rush','TDS'),
        #                                 'REC':('Rec','REC'),'YDS.1':('Rec','YDS'),'TDS.1':('Rec','TDS'),'FL':('Rush','FL')})
        #
        # df_flex[('Rec','YDS')] = df_flex[('Rec','YDS')].str.replace(',','').astype(float)
        # df_flex[('Rush','YDS')] = df_flex[('Rush','YDS')].str.replace(',','').astype(float)
        # df_flex[('Rush','TDS')] = df_flex[('Rush','TDS')].astype(float)
        # df_flex[('Rec','TDS')] = df_flex[('Rec','TDS')].astype(float)
        # df_flex[('Rec','REC')] = df_flex[('Rec','REC')].astype(float)
        # df_flex.columns = pd.MultiIndex.from_tuples(df_flex.columns)
        # df_flex.insert(5,('Rush','100 GMS'),df_flex[('Rush','YDS')].apply(compute_100_yd_games).round(1))
        # df_flex.insert(9,('Rec','100 GMS'),df_flex[('Rec','YDS')].apply(compute_100_yd_games).round(1))
        # df = pd.merge(df_qb,df_flex,on = [('','Player'),('','Team'),('','POS'),('Rush','YDS'),('Rush','TDS'),('Rush','100 GMS'),('Rush','FL')], how = 'outer').fillna(0)
        # ffp = df[('Pass','YDS')]/25 + df[('Pass','TDS')]*4+df[('Pass','INTS')]*(-2) + df[('Pass','300 GMS')]*2\
        #         + df[('Rush','YDS')]/10 + df[('Rush','TDS')]*6 + df[('Rush','FL')]*(-2) + df[('Rush','100 GMS')]*2\
        #         + df[('Rec','REC')]*0.5+ df[('Rec','YDS')]/10 + df[('Rec','TDS')]*6 + df[('Rec','100 GMS')]*2
        # df.insert(3,('','FFP'), ffp.round(1) )
        # df.insert(4,('','FPPG'),(ffp/17).round(1))
        #
        # # Get IDs with players
        # rankings = self.fp_rankings[['Player','Position','ID']].rename(columns = {'Player':('','Player'),'Position':('','POS'),'ID':('','ID')})
        # df = pd.merge(rankings,df,on = [('','Player'),('','POS')],how = 'right')
        # df.columns = pd.MultiIndex.from_tuples(df.columns)

        #
        # df_dict = {}
        # for key in ['median','high','low']:
        #     df_cur = df.loc[df[('','Team')]==key].drop(columns = ('','Team')).sort_values(('','FFP'),ascending = False).reset_index(drop = True)
        #     df_cur.index = df_cur.index + 1
        #     df_dict[key] = df_cur
        # return df_dict, df_clay
        return df_clay

    def get_roster(self,user,db, projections = None):
        roster_init = db.loc[(db['Owner']==user) & (db['Player'] != '')].reset_index()
        if projections is None:
            projections = self.projections[[('','ID'),('','FPPG')]]
            projections.columns = projections.columns.droplevel(level = 0)
        roster = pd.merge(roster_init,projections, on = 'ID',how = 'left').sort_values(by = 'FPPG',ascending = False)

        def get_first_element_or_empty_string(df, element,col):
            if len(df) > element:  # Check if the DataFrame is not empty
                return df[col].iloc[element]
            else:
                return ''

        QB1_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='QB'], 0, 'ID')
        RB1_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='RB'], 0, 'ID')
        RB2_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='RB'], 1, 'ID')
        WR1_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='WR'], 0, 'ID')
        WR2_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='WR'], 1, 'ID')
        TE1_id = get_first_element_or_empty_string(roster.loc[roster['Position']=='TE'], 0, 'ID')
        locked_ids = [QB1_id, RB1_id, RB2_id, WR1_id, WR2_id, TE1_id]
        roster_leftover = roster.loc[~roster['ID'].isin(locked_ids)]
        OP_id = get_first_element_or_empty_string(roster_leftover,0,'ID')
        locked_ids += [OP_id]
        Flex1_id = get_first_element_or_empty_string(roster.loc[(~roster['ID'].isin(locked_ids)) & (roster['Position']!= 'QB')],0,'ID')
        Flex2_id =  get_first_element_or_empty_string(roster.loc[(~roster['ID'].isin(locked_ids)) & (roster['Position']!= 'QB')],1,'ID')
        locked_ids += [Flex1_id,Flex2_id]
        pos_dict = {QB1_id:'QB1',RB1_id:'RB1',RB2_id:'RB2',WR1_id:'WR1',WR2_id:'WR2',TE1_id:'TE1',OP_id:'OP1',Flex1_id:'FLEX1',Flex2_id:'FLEX2'}
        starters = roster.loc[roster['ID'].isin(locked_ids)]
        bench = roster.loc[~roster['ID'].isin(locked_ids)]

        starters['Slot'] = starters['ID'].map(pos_dict)
        slots = ['QB1','RB1','RB2','WR1','WR2','TE1','OP1','FLEX1','FLEX2','BENCH1','BENCH2','BENCH3','BENCH4','BENCH5','BENCH6','BENCH7','BENCH8']
        bench['Slot'] = 'BENCH' + (bench.reset_index().index+1).astype(str)

        # print(starters)
        # print(bench)

        final_roster = pd.concat([starters,bench]).set_index('Slot').reindex(index = slots ).fillna('')
        final_roster = final_roster[['Position','Player','Round','Pick','Owner','Pick Type','ID','FPPG']]
        starters = final_roster.loc[~final_roster.index.str.contains('BENCH')]

        stats = starters[['Player','FPPG']]
        def convert_string_to_zero(val):
            if isinstance(val,str):
                return 0
            else:
                return val
        stats.loc['TOTAL'] = ['',stats['FPPG'].apply(convert_string_to_zero).sum()]
        return stats.loc['TOTAL','FPPG'],final_roster, stats

    def evaluate_starting_lineup(self, roster, projections = None):
        if projections is None:
            projections = self.projections[[('','ID'),('','FPPG')]]
            projections.columns = projections.columns.droplevel(level = 0)
        qb_ids = roster.loc[roster['Position'] == 'QB', 'ID']
        rb_ids = roster.loc[roster['Position'] == 'RB', 'ID']
        wr_ids = roster.loc[roster['Position'] == 'WR', 'ID']
        te_ids = roster.loc[roster['Position'] == 'TE', 'ID']
        position_ids = {
            'QB1': qb_ids.iloc[0] if not qb_ids.empty else self.replacement_ids['QB1'],
            'RB1': rb_ids.iloc[0] if not rb_ids.empty else self.replacement_ids['RB1'],
            'RB2': rb_ids.iloc[1] if len(rb_ids) > 1 else self.replacement_ids['RB2'],
            'WR1': wr_ids.iloc[0] if not wr_ids.empty else self.replacement_ids['WR1'],
            'WR2': wr_ids.iloc[1] if len(wr_ids) > 1 else self.replacement_ids['WR2'],
            'TE1': te_ids.iloc[0] if not te_ids.empty else self.replacement_ids['TE1']
        }
        # Filter and get player IDs for Flex1 and Flex2
        leftovers_flex = roster.loc[(~roster['ID'].isin(position_ids.values())) & (roster['Position'] != 'QB')]
        position_ids['Flex1'] = leftovers_flex.iloc[0]['ID'] if not leftovers_flex.empty else self.replacement_ids[
            'Flex1']
        position_ids['Flex2'] = leftovers_flex.iloc[0]['ID'] if len(leftovers_flex) > 0 else self.replacement_ids[
            'Flex2']

        # Filter and get player ID for OP position
        leftovers = roster.loc[~roster['ID'].isin(position_ids.values())]
        position_ids['OP'] = leftovers.iloc[0]['ID'] if not leftovers.empty else self.replacement_ids['OP']

        proj_pts = projections.loc[projections['ID'].isin(position_ids.values()), 'FPPG'].sum()
        return proj_pts

    def compute_VAR_odds(self, roster, player_rankings, round):
        candidates = player_rankings.reset_index()

        # Get the index of the minimum rank for each position
        min_rank_indices = candidates.groupby('Position')['Rank'].idxmin()
        candidates = candidates.loc[min_rank_indices].sort_values(by='Rank')

        # Calculate the current points and the forecast for each candidate
        cur_points = self.evaluate_starting_lineup(roster)
        def calculate_forecast(row, roster,cur_points):
            roster_new = roster.copy()
            roster_new = pd.concat([roster_new, pd.DataFrame([row])[['Position', 'Player', 'ID']]], ignore_index=True)
            forecast = self.evaluate_starting_lineup(roster_new)
            return max(forecast, cur_points)

        candidates['Forecast'] = candidates.apply(lambda row: calculate_forecast(row, roster, cur_points), axis=1)

        candidates['prob'] = np.where(candidates['Forecast'] - (candidates['Forecast'].max() - 6) > 0,
                                      candidates['Forecast'] - (candidates['Forecast'].max() - 6),
                                      0)
        candidates['prob'] /= candidates['prob'].sum()

        if candidates['Forecast'].max() - candidates['Forecast'].min() < 0.2:
            # Use additional logic
            num_players = roster.groupby('Position').size().reindex(index=['QB', 'RB', 'WR', 'TE']).fillna(0)
            num_players_ideal = pd.Series([3, 6, 6, 2], index=['QB', 'RB', 'WR', 'TE'])
            num_players_impetus = ((num_players_ideal - num_players) / num_players_ideal).clip(lower=0)
            num_players_prob = num_players_impetus / num_players_impetus.sum()

            try:
                qb_proj = self.projections.loc[
                    candidates.loc[candidates['Position'] == 'QB', ('', 'ID')], ('', 'FPPG')].values
            except KeyError:
                qb_proj = np.array([0])
            if qb_proj < 5:
                num_players_prob['QB'] *= 1 / 4
                num_players_prob /= num_players_prob.sum()

            return num_players_prob.to_frame(name='prob')
        elif round >= 7:
            num_players = roster.groupby('Position').size().reindex(index=['QB', 'RB', 'WR', 'TE']).fillna(0)
            num_players_ideal = pd.Series([1, 2, 2,0], index=['QB', 'RB', 'WR', 'TE'])
            num_players_impetus = ((num_players_ideal - num_players) / num_players_ideal).clip(lower=0)
            num_players_prob = num_players_impetus / num_players_impetus.sum()
            return num_players_prob.to_frame(name='prob')
        elif round >= 11:
            num_players = roster.groupby('Position').size().reindex(index=['QB', 'RB', 'WR', 'TE']).fillna(0)
            num_players_ideal = pd.Series([2, 2, 2,0], index=['QB', 'RB', 'WR', 'TE'])
            num_players_impetus = ((num_players_ideal - num_players) / num_players_ideal).clip(lower=0)
            num_players_prob = num_players_impetus / num_players_impetus.sum()
            return num_players_prob.to_frame(name='prob')
        elif round >= 14:
            num_players = roster.groupby('Position').size().reindex(index=['QB', 'RB', 'WR', 'TE']).fillna(0)
            num_players_ideal = pd.Series([2, 3, 3,1], index=['QB', 'RB', 'WR', 'TE'])
            num_players_impetus = ((num_players_ideal - num_players) / num_players_ideal).clip(lower=0)
            num_players_prob = num_players_impetus / num_players_impetus.sum()
            return num_players_prob.to_frame(name='prob')
        return candidates.set_index('Position')['prob'].to_frame()

        # QB_id = player_rankings.loc[]

    def trade_picks(self, trade_away = [], trade_for = []):
        trade_away_picks = self.draft_board.loc[self.draft_board.index.isin(trade_away)]
        trade_for_picks = self.draft_board.loc[self.draft_board.index.isin(trade_for)]
        if len(trade_away_picks) != len(trade_for_picks):
            print("TRADE FAILED. Must be same number of players on each side.")
            return None, None
        else:
            if ((len(trade_away_picks['Owner'].unique()) != 1) | (len(trade_for_picks['Owner'].unique()) != 1)):
                print("TRADE FAILED. Must have picks from exactly one team on each side.")
                return None, None
            self.draft_board.loc[self.draft_board.index.isin(trade_away),'Owner'] = trade_for_picks['Owner'].iloc[0]
            self.draft_board.loc[self.draft_board.index.isin(trade_for),'Owner'] = trade_away_picks['Owner'].iloc[0]
            owner_away = trade_away_picks['Owner'].iloc[0]
            owner_for = trade_for_picks['Owner'].iloc[0]
            return self.draft_board.loc[self.draft_board['Owner']==owner_away], self.draft_board.loc[self.draft_board['Owner']==owner_for]

    def reset_draft_board(self):
        self.draft_board = self.initial_draft.copy()

    def start_draft(self, user_name, simulate = False):
        random.seed(time.time())
        self.player_dict = self.reset_player_draft_rankings()
        db = self.draft_board.copy()
        user_name_list=sorted(list(db['Owner'].unique()))
        user_name_list_string = ', '.join(user_name_list)
        user_picks = db.loc[db['Owner'] == user_name]
        if not simulate:
            print("Your picks are \n\n{}\n\n".format(user_picks))
            verb = 'Starting'
        else:
            verb = 'Simulating'
        print("###############################################\n###############################################"
              "\n##### {} Draft\n###############################################\n###############################################\n".format(verb))
        prev_pick = 'None'
        prev2_pick = 'None'
        prev3_pick = 'None'
        user_name_to_use = user_name
        if simulate:
            start_time = time.time()
            user_name_to_use = ''
        index = 1
        prob_log = pd.DataFrame()
        while index <= db.index.max():
            user_name_to_use, index = self.progress_round(db, index, user_name_to_use, prob_log)
            if str(index).lower() in ['q','quit']:
                return
            index += 1
        roster = db.loc[db['Owner']==user_name].reset_index()[['Round','Pick','Slot','Position','Player','ID']].set_index('Pick')
        print("###############################################\n###############################################"
              "\n##### Draft Results\n###############################################\n###############################################\n")
        print("\n\nYour Draft:\n{}".format(roster))

        proj_points = {}
        _,final_roster,stats = self.get_roster(user_name,db)
        print("\nYour starting roster and projections.\n{}\n".format(stats))
        proj_roster = pd.DataFrame()
        full_roster_dict = {}
        for user in user_name_list:
            full_roster_dict[user] = db.loc[db['Owner']==user]
            proj_points[user],roster,_ = self.get_roster(user,db)
            roster.loc[roster.index.str.contains('BENCH'),'Player'] = '(' + roster.loc[roster.index.str.contains('BENCH'),'Position'] + ') ' \
                + roster.loc[roster.index.str.contains('BENCH'), 'Player']
            roster = roster[['Player']].rename(columns = {'Player':user})
            proj_roster = pd.concat([proj_roster,roster],axis = 1 )

        if simulate:
            print("\nSimulation time:\t{} seconds.".format(time.time()-start_time))

        proj_roster = proj_roster.reindex(columns = [owner for owner in self.pick_order if owner in proj_roster.columns])
        print("\nLeague starting lineups.\n{}".format(proj_roster))
        proj_standings = pd.DataFrame.from_dict(proj_points,orient = 'Index').rename(columns = {0:'Project FPPG'})\
            .sort_values(by = 'Project FPPG',ascending = False).reset_index().rename(columns = {'index':'Team'})
        proj_standings.index += 1
        print("\nLeague Projected Standings:\n\n{}".format(proj_standings))

        print("###############################################\n###############################################"
              "\n##### Draft Completed\n###############################################\n###############################################\n")
        entry = input("\nYou've completed the draft, would you like to save the results to csv(Y/N)?:\t")
        if entry.lower() in ['yes','y']:
            file_path = self.write_results_to_excel(db, final_roster, proj_roster, proj_standings)
            print("Mock draft results written to {}.\n".format(file_path))
        else:
            print("Mock draft results discarded.\n")
        self.prob_log = prob_log
        return db, full_roster_dict

    def write_results_to_excel(self,db,final_roster, proj_roster, proj_standings):
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime("%Y%m%d_%H%M")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        db.insert(3,'Pick',db.index)
        proj_roster.reset_index(drop = False, inplace = True)
        dataframes = [db,final_roster,proj_roster,proj_standings]
        sheet_names = ['Draft Board','My Roster','All Rosters','Projected Standings']
        for df, sheet_name in zip(dataframes,sheet_names):
            ws = wb.create_sheet(sheet_name)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Autofit column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width

                # Freeze panes
            if sheet_name == 'Draft Board':
                ws.freeze_panes = ws['B2']
                ws.auto_filter.ref = ws.dimensions
        wb.remove(ws1)
        file_path = os.path.join(INPUT_FILE_PATH,'output','MockDraft_Results_{}.xlsx'.format(formatted_time))
        wb.save(file_path)
        return file_path


    def generate_top_picks(self, db, positional_ranks, N):
        keys = ['QB','RB','WR','TE']
        df = pd.DataFrame(index = np.arange(1,N+1))
        for position in keys:
            df[(position,'Player')] = positional_ranks[position].sort_index()['Player'][:N]
            df[(position,'ID')] = positional_ranks[position].sort_index()['ID'][:N]
        df.columns = pd.MultiIndex.from_tuples(df.columns)
        return df

    def pick_cpu_player(self, player_rankings, user, db, pick, round):
        player_rankings = player_rankings.loc[~player_rankings['ID'].isin(db['ID'].unique())].sort_index()
        roster = db.loc[(db['Owner'] == user) & (db['Player'] != '')]

        # Step 1. Get historic probabilities for this pick and +- 2 picks.
        historic_pick_probs = self.historic_pick_probs.loc[pick]

        # Step 2. Get historic probabilities for the user
        historic_personal_probs = league_history.get_personal_pick_probability_by_round(self.historic_data, user, round, weight_length=5)

        # Step 3. Get probabilities by observing gaps in your rankings.
        df = player_rankings.reset_index()
        min_rank_indices = df.groupby('Position')['Rank'].idxmin()
        df = df.loc[min_rank_indices].sort_values(by='Rank')
        df['wgt'] = np.clip(np.exp((pick - df['Rank']-5)/5), 0, 1)
        df['wgt'] /= df['wgt'].sum()
        prob_gap = df[['Position', 'wgt']].set_index('Position').rename(columns={'wgt': 'rank'})
        prob_gap['historic_all'] = historic_pick_probs
        prob_gap['historic_yours'] = historic_personal_probs
        prob_gap = prob_gap.reindex(index=['QB', 'RB', 'WR', 'TE'])

        # Step 4. Team need.
        prob_gap['need'] = self.compute_VAR_odds(roster,player_rankings, round)

        # get full probs for writing to log
        prob_row_cols = pd.MultiIndex.from_product([prob_gap.columns,prob_gap.index],names = ['Prob Type','Row'])
        prob_row = pd.DataFrame(prob_gap.T.values.reshape(1,-1),columns = prob_row_cols)
        prob_row.insert(0,'User',user)
        prob_row.insert(0,'Pick',pick)

        # amplify probs by interpolated weights
        personal_weights = pd.DataFrame(
            np.array([[0.3, 0.6, 0.1, 0], [0.8, 0.2, 0, 0], [0.3, 0, 0, 0.7], [0, 0, 0, 1],[0, 0, 0, 1]]),
            columns=['rank', 'historic_all', 'historic_yours', 'need'],
            index=[0, 24, 84, 156, 204])
        def interpolate_between_rows(df, x):
            x = max(min(x, df.index.max()), df.index.min())
            lower_row = df.loc[df.index <= x].iloc[-1]
            upper_row = df.loc[df.index >= x].iloc[0]
            interpolation_factor = (x - lower_row.name) / (upper_row.name - lower_row.name)
            interpolated_vector = lower_row + interpolation_factor * (upper_row - lower_row)
            if np.isnan(interpolation_factor):
                return upper_row
            return interpolated_vector
        personal_weights = interpolate_between_rows(personal_weights, pick)

        prob_gap *= personal_weights
        probabilities = prob_gap.sum(axis= 1)

        # Early round overrides
        # QB deweighting
        qb_end_deweight = 72
        te_end_deweight = 50
        if pick <= qb_end_deweight:
            probabilities['QB'] *= (1/5)*(qb_end_deweight-pick)/qb_end_deweight + 1 *(pick/qb_end_deweight)
        if pick <= 12:
            probabilities['WR'] *= (3/5)*(6-pick)/6 + 1 *(pick/6)
        if pick <= te_end_deweight:
            probabilities['TE'] *= (3 / 5) * (te_end_deweight - pick) / te_end_deweight + 1 * (pick / te_end_deweight)
        probabilities /= probabilities.sum()

        sim = random.choices(range(len(probabilities)),probabilities)[0]
        position = probabilities.index[sim]
        dfnew = player_rankings.reset_index()
        try:
            player_id = dfnew.loc[dfnew['Position']==position].iloc[0]['ID']
        except:
            if random.random() < 0.5:
                position = 'WR'
            else:
                position = 'RB'
            player_id = dfnew.loc[dfnew['Position'] == position].iloc[0]['ID']



        # if pick > 0:
        #     print(probabilities)

        return position, player_id, prob_row


    def progress_round(self, db, index, user_name, prob_log):
        user_name_new = user_name
        index_new = index
        if db.loc[index, 'Pick Type'] == 'KEEP':
            print("Round {}, Slot {} (Pick {}):\t{} {} (KEEPER)".format(str(db.loc[index,'Round']).rjust(2),str(db.loc[index,'Slot']).rjust(2),
                                                                        str(index).rjust(3),
                                                                        db.loc[index,'Position'].rjust(2),db.loc[index,'Player'].rjust(19)))
            return user_name_new, index_new
        else:
            picker = db.loc[index, 'Owner']
            cur_rankings = self.fp_rankings.loc[~self.fp_rankings['ID'].isin(db['ID'].unique())]
            cur_pos_rankings = convert_to_position(cur_rankings)
            if picker == user_name_new:
                print("\n###############################################"
                      "\n##### Your Pick (pick {})\n###############################################\n".format(index))
                print("Your pick. Top 5 available positional players are:\n\n{}\n".format(self.generate_top_picks(db, cur_pos_rankings, 5)))
                entry = ''
                while entry not in cur_rankings['ID'].unique():
                    entry = input("Enter pick ID. Enter H for help.\t")
                    if entry.lower() in ['s','sim','simulate']:
                        print("Simulating the rest of the draft.\n\n")
                        user_name_new =  ''
                        break
                    elif entry.lower() in ['q','quit']:
                        print("Quitting draft.")
                        return user_name_new, entry
                    elif entry.lower() in ['h','help']:
                        print("\nEnter S\t\t to simulate rest of draft.\nEnter Top(N)\t to print the top N players at each position, e.g. top(12)."
                              "\nEnter R\t\t to revert to your previous draft pick, or restart if this is your first pick."
                              "\nEnter V\t\t to view your current roster."
                              "\nEnter USER\t to view the current roster of USER. Must be exact match.\n\n")
                    elif entry.lower() in ['r','rev','revert']:
                        user_picks = db.loc[(db.index < index) & (db['Owner']==user_name_new)]
                        index_new = max(1,user_picks.index.max())
                        print("\n##### Reverting back to pick {} #####\n".format(index_new))
                        db.loc[(db.index >= index_new)&(db.index < index) & (db['Pick Type'] =='PICK'),['Position','Player','ID']]=''
                        index_new -= 1
                        return user_name_new, index_new
                    elif entry.lower() in ['v','view']:
                        _, cur_roster, _ = self.get_roster(user_name,db)
                        print("\nCurrent roster:\n\n{}\n".format(cur_roster))
                    elif entry.lower() in [owner.lower() for owner in db['Owner'].unique()]:
                        owner_mapping = {owner.lower():owner for owner in db['Owner'].unique()}
                        user = owner_mapping[entry.lower()]
                        print(user)
                        _, user_roster, _ = self.get_roster(user,db)
                        print("\nRoster of {}:\n\n{}\n".format(user, user_roster))
                    elif entry.lower().startswith('top('):
                        number = int(entry.split('(',1)[1].split(')',1)[0])
                        print("\nYour pick. Top {} available positional players are:\n\n{}\n".format(number,
                            self.generate_top_picks(db, cur_pos_rankings, number)))
                    elif entry not in self.fp_rankings['ID'].unique():
                        print("\nInvalid player ID or otherwise unrecognized entry.\n\n")
                    elif entry not in cur_rankings['ID'].unique():
                        print("\nPlayer already chosen. Choose again.\n\n")
                if entry.isnumeric():
                    player_short = self.id_mapping[entry]
                    position = self.id_pos_mapping[entry]
                    print("You picked {} {}.\n".format(position, player_short))
            if picker != user_name_new:

                player_rankings = self.player_dict[picker]['ranks'].copy()
                position, entry, prob_row = self.pick_cpu_player(player_rankings, picker, db, index, db.loc[index, 'Round'])
                # player_pos_rankings = self.player_dict[picker]['pos_ranks'][position].copy()
                player_short = self.id_mapping[entry]
                prob_log = pd.concat([prob_log, prob_row])
            db.loc[index, ['Position', 'Player', 'ID']] = [position, player_short, entry]
            print("Round {}, Slot {} (Pick {}):\t{} {}".format(str(db.loc[index, 'Round']).rjust(2),
                                                               str(db.loc[index, 'Slot']).rjust(2),
                                                               str(index).rjust(3), position.rjust(2),
                                                               player_short.rjust(19)))

            return user_name_new, index_new

def main_menu():
    pass

def get_user():
    user_name_list_string = ', '.join(draft.user_name_list)
    user_name = ''
    while user_name not in draft.user_name_list:
        owner_mapping = {owner.lower(): owner for owner in draft.user_name_list}
        user_name = input("Enter user name from the following list.\n{}\nName:\t".format(user_name_list_string))
        if user_name.lower() in owner_mapping.keys():
            user_name = owner_mapping[user_name.lower()]
        else:
            print("\nPlease enter a name exactly as in the list.\n")
    print("\n###############################################\nDrafting as user:\t{}\n###############################################\n".format(user_name))
    return user_name

def startup(user_name):

    print("\n#################### HOME SCREEN ####################\n")
    print("Your options are:\n\nEnter \"Start\"\t\tto begin a mock draft."
          "\nEnter \"Simulate\"\tto simulate a mock draft."
          "\nEnter \"Trade(A,B)\"\tto change the draft order by trading picks A for picks B."
          "\n\t\t\tE.g. enter \"Trade([1,25],[12,13])\" to trade picks 1 and 25 for 12 and 13)."
          "\n\t\t\tOnly trades with the same number of picks on each side, from the same team, are supported."
          "\nEnter \"Revert\"\t\tto revert the draft order to the original, saved in \"2023 Draft\" tab of \"BKFFL Draft History\"."
          "\nEnter \"Quit\"\t\tto quit the application."
          )
    entry = input("\nYour selection:\t")
    if entry.lower() in ['start','s']:
        db = draft.start_draft(user_name = user_name, simulate = False)
        draft.db = db
    elif entry.lower() in ['simulate','sim']:
        db = draft.start_draft(user_name = user_name, simulate = True)
        draft.db = db
    elif entry.lower().startswith('trade'):
        try:
            A_str = entry.split('(',1)[1].split(']',1)[0]+']'
            B_str =  '['+ entry.split('(', 1)[1].split('[', 2)[2].split(')',1)[0]
            A =ast.literal_eval(A_str)
            B =ast.literal_eval(B_str)
            teamA,teamB = draft.trade_picks(A,B)
            if ((teamA is not None) & (teamB is not None)):
                print("\nSUCCESSFUL TRADE of picks {} FOR {}\n".format(A_str, B_str))
                print("\nTeam As new picks.\n{}\n\nTeam Bs new picks.\n{}".format(teamA,teamB))
        except:
            print("Unrecognized option. Returning null.\n")

    elif entry.lower() in ['q','quit']:
        print("Quitting application.")
        return None
    elif entry.lower() in ['revert']:
        draft.reset_draft_board()
        print("Draft board reverted to original.\n")
    else:
        print("Unrecognized option. Returning null.\n")
    startup(user_name)

if __name__ == "__main__":
    print("Welcome to Nick's BKFFL mock draft tool.\nInitializing draft setup.\n")
    draft = MockDraft()
    user_name = get_user()
    db = startup(user_name)
    # db, rosters = draft.start_draft()


